"""
data_loader.py
==============
Loads all DM test suite Excel files into a unified pandas DataFrame.
Works with the exact file/sheet/column structure found in DM_All_TestSuites.zip.

Usage:
    from data_loader import load_all_suites
    df = load_all_suites("path/to/DM_All_TestSuites/")
    print(df.shape)
    print(df.columns.tolist())
"""

import os
import re
import pandas as pd
from openpyxl import load_workbook
from typing import Optional


# ── File manifest ────────────────────────────────────────────────────────────
# Maps filename → (sheet_name, suite_code, format_type)
# format_type: "step"  = Testcase Name / Step / Description columns
#              "rich"  = has Test_Type, Test_Priority, Test_Result columns
#              "ops"   = DM_Tab_Operations layout
#              "setup" = DM_Product_Setup layout

FILE_MANIFEST = {
    "DM_SANITY_v2.xlsx":                              ("Sanity_TestCases",         "SANITY",       "step"),
    "Automated_Data_Management_System_ SanityTestProcedure_v5.xlsx":
                                                       ("System_Sanity_TestCases",  "SYS_SANITY",   "step"),
    "DM_CTR_v1.xlsx":                                 ("CTR_TestCases",             "CTR",          "step"),
    "DM_FRM_v1.xlsx":                                 ("FRM_TestCases",             "FRM",          "step"),
    "DM_GGP_F_Test_Cases_v1.xlsx":                    ("GGP_F_Test_Cases",          "GGP_F",        "step"),
    "DM_GGP_S_Test_Cases_v1.xlsx":                    ("DM_GGP_Test_Cases",         "GGP_S",        "step"),
    "DM_GPN_v1.xlsx":                                 ("GPN_TestCases",             "GPN",          "step"),
    "DM_PDT_v1.xlsx":                                 ("PDT_Test_Cases",            "PDT",          "step"),
    "Defect_Automation_Test_Definitions_v1.xlsx":     ("Defect_Automation",         "DEFECT",       "step"),
    "DM_PFD_Test_Cases_V38.xlsx":                     ("DM_PFD_Test_Cases",         "PFD",          "rich"),
    "DM_TSK_AsApplied_Test_Cases_V7.0.xlsx":          ("DM_TSK_Test_Cases",         "TSK",          "rich_exec"),
    "DM_ShapeFile_Test_Cases_V1.1.xlsx":              ("SHP_Test_Cases",            "SHP",          "rich"),
    "DM_Operation_Test_Cases_V3.0.xlsx":              ("DM_Tab_Operations",         "OPS",          "ops"),
    "DM_Product_Application_Setup_0.3.xlsx":          ("DM_Product_Setup",          "PROD_SETUP",   "setup"),
}


def _clean(val) -> Optional[str]:
    """Normalize a cell value to clean string or None."""
    if val is None:
        return None
    s = str(val).strip()
    return None if s in ("", "None", "NA", "N/A", "nan") else s


def _load_step_format(path: str, sheet: str, suite: str) -> list[dict]:
    """
    Parse files where each test case spans multiple rows (one per step).
    Columns: Testcase Name | Description | Step | Step Description | Expected | Linked WI | Type
    Returns one record per test case with step_count computed.
    """
    wb = load_workbook(path, read_only=True)
    ws = wb[sheet]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    records = []
    current = None

    for r in rows[1:]:  # skip header
        tc_name = _clean(r[0])
        description = _clean(r[1])
        step_num = r[2]
        step_desc = _clean(r[3]) if len(r) > 3 else None
        linked_wi = _clean(r[5]) if len(r) > 5 else None
        tc_type = _clean(r[6]) if len(r) > 6 else None

        if tc_name is not None:
            if current:
                records.append(current)
            current = {
                "test_id":        tc_name,
                "suite":          suite,
                "description":    description or "",
                "step_count":     1 if step_num else 0,
                "linked_work_item": linked_wi,
                "test_type":      tc_type,
                "priority":       None,
                "test_level":     None,
                "result":         None,
                "script_result":  None,
                "defect_id":      None,
                "has_script":     False,
                "script_name":    None,
                "build_number":   None,
                "is_automated":   tc_type == "Automated" if tc_type else None,
                "is_defect_test": suite == "DEFECT",
            }
        elif current and step_num is not None:
            current["step_count"] = (current["step_count"] or 0) + 1

    if current:
        records.append(current)

    return records


def _load_rich_format(path: str, sheet: str, suite: str) -> list[dict]:
    """
    Parse files with full column set: Test_Type, Test_Priority, Test_Level,
    Test_Script_Name, Script created?, Comment.
    Used by PFD and SHP.
    """
    wb = load_workbook(path, read_only=True)
    ws = wb[sheet]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    header = rows[0]
    col = {str(h): i for i, h in enumerate(header) if h is not None}

    records = []
    for r in rows[1:]:
        def g(name):
            return _clean(r[col[name]]) if name in col and col[name] < len(r) else None

        tc_id = g("Test_Case_ID")
        if not tc_id:
            continue

        script_name = g("Test_Script_Name")
        has_script = script_name is not None

        comment = g("Comment") or ""
        defect_in_comment = None
        # Extract defect IDs embedded in Comment column (e.g. "pfhmi-dev-defects-12345")
        dm = re.findall(r'(?:eagle-systems|pfhmi-dev-defects|MADSW)-[\w-]+', comment)
        if dm:
            defect_in_comment = dm[0]

        records.append({
            "test_id":         tc_id,
            "suite":           suite,
            "description":     g("Test_Case_Description ") or "",
            "step_count":      None,
            "linked_work_item": None,
            "test_type":       g("Test_Type"),
            "priority":        g("Test_Priority"),
            "test_level":      g("Test_Level"),
            "result":          None,
            "script_result":   None,
            "defect_id":       defect_in_comment,
            "has_script":      has_script,
            "script_name":     script_name,
            "build_number":    None,
            "is_automated":    has_script,
            "is_defect_test":  False,
        })

    return records


def _load_rich_exec_format(path: str, sheet: str, suite: str) -> list[dict]:
    """
    Parse TSK file which has both test-design AND execution result columns:
    Test_Result, Defect_ID, Build_Number, Script_Test_Result, Actual_Result.
    This is the gold standard format for ML training.
    """
    wb = load_workbook(path, read_only=True)
    ws = wb[sheet]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    header = rows[0]
    col = {str(h): i for i, h in enumerate(header) if h is not None}

    records = []
    for r in rows[1:]:
        def g(name):
            return _clean(r[col[name]]) if name in col and col[name] < len(r) else None

        tc_id = g("Test_Case_ID")
        if not tc_id:
            continue

        defect = g("Defect_ID")

        records.append({
            "test_id":         tc_id,
            "suite":           suite,
            "description":     g("Test_Case_Description ") or "",
            "step_count":      None,
            "linked_work_item": None,
            "test_type":       g("Test_Type"),
            "priority":        g("Test_Priority"),
            "test_level":      g("Test_Level"),
            "result":          g("Test_Result"),
            "script_result":   g("Script_Test_Result"),
            "defect_id":       defect if (defect and defect not in ("NA",)) else None,
            "has_script":      g("Test_Script_Name") is not None,
            "script_name":     g("Test_Script_Name"),
            "build_number":    g("Build_Number"),
            "is_automated":    g("Test_Type") == "Automated" if g("Test_Type") else None,
            "is_defect_test":  False,
        })

    return records


def _load_ops_format(path: str, sheet: str, suite: str) -> list[dict]:
    wb = load_workbook(path, read_only=True)
    ws = wb[sheet]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    header = rows[0]
    col = {str(h): i for i, h in enumerate(header) if h is not None}

    records = []
    for r in rows[1:]:
        def g(name):
            return _clean(r[col[name]]) if name in col and col[name] < len(r) else None

        tc_id = g("Test_Case_ID")
        if not tc_id:
            continue

        records.append({
            "test_id": tc_id, "suite": suite,
            "description": g("Test_Case_Description ") or "",
            "step_count": None, "linked_work_item": None,
            "test_type": None, "priority": None, "test_level": None,
            "result": g("Result"), "script_result": None, "defect_id": None,
            "has_script": None, "script_name": None, "build_number": None,
            "is_automated": None, "is_defect_test": False,
        })

    return records


def _load_setup_format(path: str, sheet: str, suite: str) -> list[dict]:
    wb = load_workbook(path, read_only=True)
    ws = wb[sheet]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    header = rows[0]
    col = {str(h): i for i, h in enumerate(header) if h is not None}

    records = []
    for r in rows[1:]:
        def g(name):
            return _clean(r[col[name]]) if name in col and col[name] < len(r) else None

        tc_id = g("Test Case ID")
        if not tc_id:
            continue

        defect = g("Defect ID") or g("Defect_ID")

        records.append({
            "test_id": tc_id, "suite": suite,
            "description": g("Test Case Description ") or "",
            "step_count": None, "linked_work_item": None,
            "test_type": g("Test_Type"), "priority": None, "test_level": None,
            "result": g("Result"), "script_result": None,
            "defect_id": defect if (defect and defect not in ("NA",)) else None,
            "has_script": None, "script_name": None, "build_number": None,
            "is_automated": None, "is_defect_test": False,
        })

    return records


LOADERS = {
    "step":       _load_step_format,
    "rich":       _load_rich_format,
    "rich_exec":  _load_rich_exec_format,
    "ops":        _load_ops_format,
    "setup":      _load_setup_format,
}


def load_all_suites(folder: str) -> pd.DataFrame:
    """
    Load all test suite Excel files from the given folder.
    Returns a unified DataFrame with consistent columns.

    Parameters
    ----------
    folder : str
        Path to directory containing the xlsx files.

    Returns
    -------
    pd.DataFrame with columns:
        test_id, suite, description, step_count, linked_work_item,
        test_type, priority, test_level, result, script_result,
        defect_id, has_script, script_name, build_number,
        is_automated, is_defect_test
    """
    all_records = []

    for filename, (sheet, suite, fmt) in FILE_MANIFEST.items():
        path = os.path.join(folder, filename)
        if not os.path.exists(path):
            print(f"  [skip] {filename} not found")
            continue
        try:
            loader = LOADERS[fmt]
            records = loader(path, sheet, suite)
            all_records.extend(records)
            print(f"  [ok] {suite:<12} {len(records):>4} tests  ({filename})")
        except Exception as e:
            print(f"  [err] {filename}: {e}")

    df = pd.DataFrame(all_records)

    # Normalize priority column
    df["priority"] = df["priority"].str.strip().str.title()

    # Derive module from test_id prefix (e.g. TC_DM_PFD_001 → PFD)
    df["module"] = df["test_id"].apply(_extract_module)

    # Derive test_number (integer ordering within module)
    df["test_number"] = df["test_id"].apply(_extract_number)

    # Mark tests that have any defect linkage
    df["has_defect"] = df["defect_id"].notna()

    # Normalize result
    df["result_clean"] = df["result"].str.strip().str.title() if df["result"].notna().any() else df["result"]

    print(f"\n  Total: {len(df)} test cases across {df['suite'].nunique()} suites")
    return df


def _extract_module(test_id: str) -> str:
    """TC_DM_PFD_001 → PFD,  MADSW-12345 → MADSW,  TC_DM_Sanity_001 → SANITY"""
    if not isinstance(test_id, str):
        return "UNKNOWN"
    m = re.match(r'TC_DM_([A-Za-z]+)_', test_id)
    if m:
        return m.group(1).upper()
    m2 = re.match(r'(MADSW)-', test_id)
    if m2:
        return "MADSW"
    m3 = re.match(r'test_TC_DM_([A-Za-z]+)_', test_id)
    if m3:
        return m3.group(1).upper()
    return "OTHER"


def _extract_number(test_id: str) -> Optional[int]:
    """TC_DM_PFD_042 → 42"""
    if not isinstance(test_id, str):
        return None
    m = re.search(r'_(\d+)$', test_id)
    return int(m.group(1)) if m else None


if __name__ == "__main__":
    import sys
    folder = sys.argv[1] if len(sys.argv) > 1 else "/tmp/DM_All_TestSuites"
    print(f"Loading from: {folder}\n")
    df = load_all_suites(folder)
    print("\nColumn dtypes:")
    print(df.dtypes)
    print("\nSample:")
    print(df[["test_id", "suite", "module", "priority", "test_type", "has_script", "has_defect", "result"]].head(10))
    print(f"\nPriority value counts:\n{df['priority'].value_counts(dropna=False)}")
    print(f"\nModule value counts:\n{df['module'].value_counts()}")
