"""
run_pipeline.py
================
Main entry point for the DM ML testing pipeline.

Loads all test suites → engineers features → runs ML models →
outputs a ranked test execution order + Excel report.

Usage:
    python run_pipeline.py --data /path/to/DM_All_TestSuites/ --output ./outputs/

What this script produces:
  outputs/
    ml_test_report.xlsx         ← main deliverable (send to your manager)
    prioritized_test_order.csv  ← feed into TestRunner.py
    models/prioritizer.pkl      ← saved model
    models/anomaly_detector.pkl ← saved anomaly model

Connect to TestRunner.py:
    In your TestRunner.py, replace the default test execution order with:
        import pandas as pd
        ranked = pd.read_csv("outputs/prioritized_test_order.csv")
        test_order = ranked["test_id"].tolist()
"""

import os
import sys
import argparse
from datetime import datetime
from typing import Optional

import pandas as pd
import numpy as np

sys.path.insert(0, os.path.dirname(__file__))
from data_loader import load_all_suites
from feature_engineering import build_static_features, get_feature_matrix
from ml_models import TestPrioritizer, BenchmarkAnomalyDetector, FlakyTestDetector


def parse_args():
    parser = argparse.ArgumentParser(description="DM ML Testing Pipeline")
    parser.add_argument("--data", default="/tmp/DM_All_TestSuites",
                        help="Path to folder with DM_All_TestSuites xlsx files")
    parser.add_argument("--output", default="./outputs",
                        help="Output directory")
    parser.add_argument("--benchmark", default=None,
                        help="Path to benchmark CSV (optional)")
    parser.add_argument("--history", default=None,
                        help="Path to execution history CSV (optional)")
    return parser.parse_args()


def generate_excel_report(
    ranked_df: pd.DataFrame,
    flaky_df: pd.DataFrame,
    anomaly_df: Optional[pd.DataFrame],
    suite_stats: pd.DataFrame,
    output_path: str,
    mode: str = "rule_based",
):
    """Generate a multi-sheet Excel report with ML insights."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import (
            PatternFill, Font, Alignment, Border, Side
        )
        from openpyxl.utils.dataframe import dataframe_to_rows
        from openpyxl.chart import BarChart, Reference
    except ImportError:
        print("[Report] openpyxl not available — saving CSV instead")
        ranked_df.to_csv(output_path.replace(".xlsx", ".csv"), index=False)
        return

    wb = Workbook()

    # ── Color palette ────────────────────────────────────────────────────
    COLORS = {
        "Critical": "C0392B",
        "High":     "E74C3C",
        "Medium":   "F39C12",
        "Low":      "27AE60",
        "header_bg":"2C3E50",
        "header_fg":"FFFFFF",
        "row_alt":  "F8F9FA",
    }

    def header_style(cell):
        cell.fill = PatternFill("solid", fgColor=COLORS["header_bg"])
        cell.font = Font(color=COLORS["header_fg"], bold=True, size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    def risk_fill(label: str):
        color = COLORS.get(str(label), "FFFFFF")
        return PatternFill("solid", fgColor=color)

    def border_thin():
        s = Side(style="thin", color="D0D0D0")
        return Border(left=s, right=s, top=s, bottom=s)

    # ── Sheet 1: Prioritized Test Order ──────────────────────────────────
    ws1 = wb.active
    ws1.title = "Prioritized Test Order"

    cols_to_show = ["rank", "test_id", "suite", "priority", "test_type",
                    "risk_label", "risk_pct", "has_defect_link", "description"]
    export_cols = [c for c in cols_to_show if c in ranked_df.columns]
    display = ranked_df[export_cols].head(200)  # max 200 rows in report

    # Write header
    for col_idx, col_name in enumerate(export_cols, 1):
        cell = ws1.cell(row=1, column=col_idx, value=col_name.replace("_", " ").title())
        header_style(cell)
    ws1.freeze_panes = "A2"

    # Write data rows
    for row_idx, (_, row) in enumerate(display.iterrows(), 2):
        alt = row_idx % 2 == 0
        for col_idx, col_name in enumerate(export_cols, 1):
            cell = ws1.cell(row=row_idx, column=col_idx, value=row.get(col_name, ""))
            cell.border = border_thin()
            if col_name == "risk_label":
                cell.fill = risk_fill(str(row.get("risk_label", "")))
                cell.font = Font(color="FFFFFF", bold=True, size=9)
                cell.alignment = Alignment(horizontal="center")
            elif alt:
                cell.fill = PatternFill("solid", fgColor=COLORS["row_alt"])

    # Column widths
    col_widths = {"rank": 6, "test_id": 25, "suite": 12, "priority": 10,
                  "test_type": 14, "risk_label": 10, "risk_pct": 10,
                  "has_defect_link": 12, "description": 60}
    for col_idx, col_name in enumerate(export_cols, 1):
        ws1.column_dimensions[
            ws1.cell(1, col_idx).column_letter
        ].width = col_widths.get(col_name, 15)

    # ── Sheet 2: Suite Risk Summary ───────────────────────────────────────
    ws2 = wb.create_sheet("Suite Risk Summary")
    suite_cols = ["suite", "test_count", "critical_count", "high_count",
                  "medium_count", "low_count", "avg_risk_pct"]
    suite_display = [c for c in suite_cols if c in suite_stats.columns]

    for col_idx, col_name in enumerate(suite_display, 1):
        cell = ws2.cell(row=1, column=col_idx,
                        value=col_name.replace("_", " ").title())
        header_style(cell)

    for row_idx, (_, row) in enumerate(suite_stats.iterrows(), 2):
        for col_idx, col_name in enumerate(suite_display, 1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=row.get(col_name, 0))
            cell.border = border_thin()
            if row_idx % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=COLORS["row_alt"])

    for col_idx in range(1, len(suite_display) + 1):
        ws2.column_dimensions[ws2.cell(1, col_idx).column_letter].width = 18

    # ── Sheet 3: Flaky Test Candidates ───────────────────────────────────
    ws3 = wb.create_sheet("Flaky Test Candidates")
    flaky_cols = ["test_id", "suite", "flakiness_keyword_count",
                  "heuristic_flakiness_score", "flakiness_risk", "description"]
    flaky_display = [c for c in flaky_cols if c in flaky_df.columns]
    flaky_top = flaky_df[flaky_df["heuristic_flakiness_score"] > 0].head(50)

    for col_idx, col_name in enumerate(flaky_display, 1):
        cell = ws3.cell(row=1, column=col_idx,
                        value=col_name.replace("_", " ").title())
        header_style(cell)

    for row_idx, (_, row) in enumerate(flaky_top.iterrows(), 2):
        for col_idx, col_name in enumerate(flaky_display, 1):
            ws3.cell(row=row_idx, column=col_idx, value=row.get(col_name, ""))

    ws3.column_dimensions["A"].width = 25
    ws3.column_dimensions["E"].width = 12
    ws3.column_dimensions["F"].width = 60

    # ── Sheet 4: Anomaly Detection (if available) ────────────────────────
    if anomaly_df is not None:
        ws4 = wb.create_sheet("Benchmark Anomalies")
        # Flatten any list columns to strings for Excel serialization
        adf = anomaly_df.copy()
        for col in adf.columns:
            if adf[col].apply(lambda x: isinstance(x, list)).any():
                adf[col] = adf[col].apply(lambda x: ", ".join(x) if isinstance(x, list) else x)
        for col_idx, col_name in enumerate(adf.columns, 1):
            cell = ws4.cell(row=1, column=col_idx,
                            value=col_name.replace("_", " ").title())
            header_style(cell)
        for row_idx, (_, row) in enumerate(adf.iterrows(), 2):
            for col_idx, (col_name, val) in enumerate(row.items(), 1):
                cell = ws4.cell(row=row_idx, column=col_idx, value=val)
                if col_name == "severity" and val == "Critical":
                    cell.fill = PatternFill("solid", fgColor=COLORS["Critical"])
                    cell.font = Font(color="FFFFFF", bold=True)

    # ── Sheet 5: ML Info ─────────────────────────────────────────────────
    ws5 = wb.create_sheet("ML Model Info")
    info = [
        ("Generated", datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("Mode", mode),
        ("Total tests analyzed", len(ranked_df)),
        ("Tool", "CNH DM ML Pipeline v1.0"),
        ("Models used", "Random Forest (prioritization), Isolation Forest (anomaly detection)"),
        ("Feature count", "30+ engineered features from test metadata"),
        ("", ""),
        ("How to use", ""),
        ("1.", "Use 'Prioritized Test Order' tab for test execution sequence"),
        ("2.", "Run Critical/High risk tests FIRST in regression cycle"),
        ("3.", "Monitor 'Flaky Test Candidates' for infrastructure issues"),
        ("4.", "Check 'Benchmark Anomalies' after each build"),
        ("", ""),
        ("Risk Labels", ""),
        ("Critical", "Run first — highest failure probability"),
        ("High", "Run in first quartile of regression"),
        ("Medium", "Standard regression order"),
        ("Low", "Run last — minimal risk"),
    ]
    ws5.column_dimensions["A"].width = 25
    ws5.column_dimensions["B"].width = 60
    for row_idx, (key, val) in enumerate(info, 1):
        ws5.cell(row=row_idx, column=1, value=key)
        ws5.cell(row=row_idx, column=2, value=val)

    wb.save(output_path)
    print(f"[Report] Saved Excel report: {output_path}")


def compute_suite_stats(ranked_df: pd.DataFrame) -> pd.DataFrame:
    """Compute per-suite risk statistics."""
    stats = []
    for suite, group in ranked_df.groupby("suite"):
        label_counts = group["risk_label"].value_counts()
        stats.append({
            "suite": suite,
            "test_count": len(group),
            "critical_count": label_counts.get("Critical", 0),
            "high_count": label_counts.get("High", 0),
            "medium_count": label_counts.get("Medium", 0),
            "low_count": label_counts.get("Low", 0),
            "avg_risk_pct": round(group["risk_pct"].mean(), 1),
        })
    return pd.DataFrame(stats).sort_values("avg_risk_pct", ascending=False)


def main():
    args = parse_args()
    os.makedirs(args.output, exist_ok=True)
    os.makedirs(os.path.join(args.output, "models"), exist_ok=True)

    print("=" * 60)
    print("CNH Data Management — ML Testing Pipeline")
    print(f"Started: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    print("=" * 60)

    # ── Step 1: Load data ────────────────────────────────────────────────
    print("\n[1/5] Loading test suites...")
    df = load_all_suites(args.data)

    # ── Step 2: Feature engineering ──────────────────────────────────────
    print("\n[2/5] Engineering features...")
    feat = build_static_features(df)

    # If execution history provided, add richer features
    exec_history = None
    if args.history and os.path.exists(args.history):
        from feature_engineering import build_execution_features
        print(f"  Loading execution history from {args.history}...")
        exec_history = pd.read_csv(args.history, parse_dates=["run_date"])
        feat = build_execution_features(feat, exec_history)
        use_exec = True
    else:
        use_exec = False
        print("  No execution history — using static features only")
        print("  (Feed TestRunner.py results back to improve ML accuracy)")

    X, y, feature_names = get_feature_matrix(feat, use_execution=use_exec)
    print(f"  Feature matrix: {X.shape[0]} tests × {X.shape[1]} features")
    print(f"  Proxy labels: {y.sum()} risky ({y.mean()*100:.1f}%), "
          f"{(1-y).sum()} safe")

    # ── Step 3: Test prioritization ───────────────────────────────────────
    print("\n[3/5] Running test prioritization model...")
    prioritizer = TestPrioritizer()

    if y.sum() >= 10:
        prioritizer.fit_ml(X, y, feature_names)
        mode = "ml"
    else:
        print("  (Insufficient labels for ML — using rule-based ranking)")
        prioritizer.fit_rule_based(feat)
        mode = "rule_based"

    ranked = prioritizer.rank(feat, X if mode == "ml" else None)
    prioritizer.save(os.path.join(args.output, "models", "prioritizer.pkl"))

    # ── Step 4: Flaky test detection ──────────────────────────────────────
    print("\n[4/5] Detecting flaky test candidates...")
    flaky_detector = FlakyTestDetector()

    if exec_history is not None:
        flaky_df = flaky_detector.detect_from_history(exec_history)
        flaky_df = flaky_df.merge(feat[["test_id", "suite", "description"]],
                                  on="test_id", how="left")
    else:
        flaky_df = flaky_detector.detect_heuristic(feat)

    # ── Step 5: Benchmark anomaly detection ───────────────────────────────
    print("\n[5/5] Benchmark anomaly detection...")
    anomaly_df = None

    if args.benchmark and os.path.exists(args.benchmark):
        bench = pd.read_csv(args.benchmark)
        metric_cols = [c for c in bench.columns
                       if any(t in c.lower() for t in ["time", "ms", "sec", "duration"])
                       and c not in ("build_id", "timestamp")]
        if metric_cols:
            print(f"  Detected metrics: {metric_cols}")
            detector = BenchmarkAnomalyDetector()
            # Use first 80% as training, last 20% as detection
            split = int(len(bench) * 0.8)
            if split >= 5:
                detector.fit(bench.iloc[:split], metric_cols)
                anomaly_df = detector.detect(bench.iloc[split:])
                detector.save(os.path.join(args.output, "models", "anomaly_detector.pkl"))
            else:
                print("  Need at least 5 benchmark builds — skipping anomaly detection")
        else:
            print("  No timing columns detected in benchmark file")
    else:
        print("  No benchmark file provided — skipping")
        print("  (Provide --benchmark path/to/benchmarks.csv to enable)")

    # ── Output: CSV for TestRunner.py ────────────────────────────────────
    csv_path = os.path.join(args.output, "prioritized_test_order.csv")
    ranked[["rank", "test_id", "suite", "risk_label", "risk_pct",
            "description"]].to_csv(csv_path, index=False)
    print(f"\n[Output] TestRunner order saved: {csv_path}")

    # ── Output: Excel report ─────────────────────────────────────────────
    suite_stats = compute_suite_stats(ranked)
    excel_path = os.path.join(args.output, "ml_test_report.xlsx")
    generate_excel_report(ranked, flaky_df, anomaly_df, suite_stats, excel_path, mode)

    # ── Summary ───────────────────────────────────────────────────────────
    print("\n" + "=" * 60)
    print("PIPELINE COMPLETE")
    print("=" * 60)
    print(f"\nTests analyzed: {len(ranked)}")
    label_counts = ranked["risk_label"].value_counts()
    for label in ["Critical", "High", "Medium", "Low"]:
        count = label_counts.get(label, 0)
        print(f"  {label:10}: {count:4} tests ({count/len(ranked)*100:.1f}%)")

    print("\nTop 10 highest-risk tests (run these FIRST):")
    print(ranked[["rank", "test_id", "suite", "risk_label"]].head(10).to_string(index=False))

    print(f"\nOutputs in: {args.output}")
    print(f"  {csv_path}")
    print(f"  {excel_path}")


if __name__ == "__main__":
    main()
